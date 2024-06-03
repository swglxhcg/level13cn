import openpyxl

# 读取文本文件a
with open('UpgradeData.json', 'r', encoding='utf-8') as f:
    lines = f.readlines()

print("OK_a")
# 读取Excel文件b
wb = openpyxl.load_workbook('UpgradeDataname.xlsx')
ws = wb.active
print("OK_b")
# 创建一个字典，用于存储Excel文件中第一列和第二列的文本
replace_dict = {}
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
    replace_dict[row[0].value] = row[1].value

# 遍历文本文件的每一行，替换文本
new_lines = []
for line in lines:
    for key, value in replace_dict.items():
        line = line.replace(key, value)
    new_lines.append(line)

# 将处理后的文本写入新的文本文件
with open('c.json', 'w', encoding='utf-8') as f:
    f.writelines(new_lines)
